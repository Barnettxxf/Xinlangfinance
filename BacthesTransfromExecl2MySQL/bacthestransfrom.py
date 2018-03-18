# -*- coding:utf-8 -*-
'''
author: Barnett
'''
import time
import handler_MySQL, config_MySQL
import openpyxl
import pandas as pd
import os
import re


class Excel2Mysql():
    """
    Save excel data in target file path as MySQL table.
    """
    def __init__(self, filepath=None):
        self.mysql = handler_MySQL.HandlerMysql()
        self.path = filepath

    def transform(self):
        filenames = self._getexcelfile()
        if len(filenames) != 0:
            for filename in filenames:
                print('File: ', filename)
                self._logs(filename)
                sheetnames = self._getsheetnames(filename)
                for sheetname in sheetnames:
                    print('Sheet: ', sheetname)
                    data = self._loaddata(filename, sheetname)
                    sql_insert = '''
                    INSERT INTO xinlangdata(板块,公司家数,平均价格,总涨跌额,总涨跌幅,总成交量,
                    总成交额,领涨股,领涨股代码,领涨股涨跌幅,领涨股当前价,领涨股涨跌额,板块链接,领涨股链接,日期,分类) VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
                    '''
                    self.mysql.insertmany(sql_insert, data.values.tolist())
                    self._commit()
            self._close()
        else:
            print('There is no any new dataset.')

    def _getsheetnames(self, excelfile):
        wb = openpyxl.load_workbook(excelfile)
        sheetnames = wb.sheetnames
        if 'Sheet' in sheetnames:
            sheetnames.remove('Sheet')
        return sheetnames

    def _getexcelfile(self):
        self.__chdir()
        filenames = os.listdir(os.curdir)
        excel_files = []
        for each in filenames:
            sign = each.split('.')[-1]
            if sign == 'xlsx' and "~" not in each:
                excel_files.append(each)
        if 'log.txt' in filenames:
            new_files = self._filter(excel_files)
        else:
            new_files = excel_files
        return new_files

    def _loaddata(self, filename, sheetname):
        data_time = re.search('[0-9]+\-[0-9]+\-[0-9]+', filename).group(0)
        data = pd.read_excel(filename, sheetname=sheetname)
        data = data.dropna(axis=0)
        try:
            data['涨跌幅　'] = data['涨跌幅　'].apply(lambda x: float(x[:-2]) * .01)
            data['涨跌幅'] = data['涨跌幅'].apply(lambda x: float(x[:-2]) * .01)
        except TypeError:
            pass
        data['日期'] = data_time
        data['分类'] = sheetname
        return data

    def _commit(self):
        self.mysql.commit()

    def _close(self):
        self.mysql.close()

    def __chdir(self):
        os.chdir(config_MySQL.TARGETDIR)

    def _logs(self, filename):
        with open('log.txt', 'a') as f:
            f.write(filename+'\n')

    def _filter(self, filenames):
        with open('log.txt', 'r') as f:
            filter_condition = f.readlines()
        new_files = []
        for each in filenames:
            if (each+'\n') not in filter_condition:
                new_files.append(each)
        return new_files


if __name__ == '__main__':
    start_time = time.time()
    run = Excel2Mysql()
    run.transform()
    print('{} secs'.format(time.time()-start_time))
