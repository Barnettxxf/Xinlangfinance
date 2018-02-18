# -*- coding=utf-8 -*-
import openpyxl
from datetime import datetime


class Outputer(object):
        def __init__(self):
                self.wb = None
                self.ws = None
                now_time = datetime.now().strftime('%Y-%m-%d')
                self.filename = '新浪财经%s' % now_time

        def creat_outputer(self):
                print('create  file...')
                self.wb = openpyxl.Workbook()
                self.wb.save('%s.xlsx' % self.filename)
                print('creat_outputer:', self.filename)

        def outputer(self):
                pass

        def collect_data(self, new_data):
                title = new_data.pop()
                print('collect_data:', self.filename)
                self.wb = openpyxl.load_workbook('%s.xlsx' % self.filename)
                try:
                        self.ws = self.wb.get_sheet_by_name('%s' % title)
                except KeyError:
                        self.ws = self.wb.create_sheet('%s' % title)
                        list_title = [
                            '板块',
                            '公司家数',
                            '平均价格',
                            '涨跌额',
                            '涨跌幅',
                            '总成交量(手)',
                            '总成交额(万元)',
                            '领涨股',
                            '领涨股代码',
                            '涨跌幅　',
                            '当前价',
                            '涨跌额',
                            '板块链接',
                            '领涨股链接']
                        self.ws.append(list_title)

                self.ws.append(new_data)
                self.wb.save('%s.xlsx' % self.filename)
